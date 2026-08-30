from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import threading
import time

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from kongpu_api.generator import generate_bundle
import kongpu_api.main as api_main
from kongpu_api.models import ProgramBranch, ProgramCommit, ProgramWorkspace
from kongpu_api.repository import RepositoryError, safe_file


def _lock_revision(client: TestClient, revision: dict) -> dict:
    for issue in revision["issues"]:
        if issue["severity"] == "warning":
            response = client.post(
                f"/api/v1/spec-revisions/{revision['id']}/warnings/{issue['id']}/accept",
                json={"reason": "自动化回归使用范例规格", "expected_revision": revision["revision"]},
            )
            assert response.status_code == 200, response.text
            revision = response.json()
    for view in revision["required_views"]:
        response = client.put(
            f"/api/v1/spec-revisions/{revision['id']}/confirmations/{view}",
            json={"confirmed_by": "自动化测试", "expected_revision": revision["revision"]},
        )
        assert response.status_code == 200, response.text
        revision = response.json()
    response = client.post(
        f"/api/v1/spec-revisions/{revision['id']}/lock",
        json={"confirmed_by": "自动化测试", "expected_revision": revision["revision"]},
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_inovance_h5u_profile_template_generation_audit_and_reference_simulation(
    client: TestClient,
) -> None:
    created = client.post(
        "/api/v1/projects",
        json={
            "name": "H5U 自动化回归线",
            "plc_brand": "汇川技术",
            "plc_series": "H5U",
            "plc_model": "H5U-1614MTD-A8",
        },
    )
    assert created.status_code == 201, created.text
    project = created.json()
    assert (project["plc_brand"], project["plc_series"], project["plc_model"]) == ("汇川技术", "H5U", "H5U-1614MTD-A8")

    template = client.post(f"/api/v1/projects/{project['id']}/templates?kind=example")
    assert template.status_code == 200, template.text
    workbook = load_workbook(BytesIO(template.content), read_only=True, data_only=True)
    meta = {row[0]: row[1] for row in workbook["_meta"].iter_rows(values_only=True) if row and row[0]}
    project_row = next(workbook["Project"].iter_rows(min_row=2, values_only=True))
    assert meta["plc_brand"] == "汇川技术"
    assert meta["plc_series"] == "H5U"
    assert project_row[3:6] == ("汇川技术", "H5U", "H5U-1614MTD-A8")

    imported = client.post(
        f"/api/v1/projects/{project['id']}/imports",
        files={"file": ("h5u.xlsx", template.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert imported.status_code == 201, imported.text
    locked = _lock_revision(client, imported.json()["revision"])
    generated = client.post(
        f"/api/v1/projects/{project['id']}/generation-runs",
        json={"spec_revision_id": locked["revision"]["id"], "branch_name": "generated/h5u-profile"},
    )
    assert generated.status_code == 201, generated.text
    run = generated.json()
    control_ir = next(item for item in run["artifacts"] if item["path"] == "generated/ControlIR.json")
    assert control_ir["content_hash"]

    files = client.get(f"/api/v1/branches/{run['branch_id']}/files").json()
    gvl = client.get(f"/api/v1/branches/{run['branch_id']}/files/src/GVL_IO.st").json()["content"]
    readme = client.get(f"/api/v1/branches/{run['branch_id']}/files/README.md").json()["content"]
    assert files["branch"]["status"] == "clean"
    assert "logical address X010" in gvl
    assert " AT %X010" not in gvl
    assert "inovance-h5u-st-v1" in readme
    assert "AutoShop" in readme

    audit = client.post(f"/api/v1/generation-runs/{run['id']}/audit")
    assert audit.status_code == 200, audit.text
    assert audit.json()["status"] == "review_ready"
    review = client.get(f"/api/v1/projects/{project['id']}/automated-reviews").json()[0]
    assert review["status"] == "passed"
    assert {item["id"] for item in review["external_validation_gates"]} >= {"autoshop_compile", "autoshop_simulation", "h5u_hardware_validation"}

    simulation = client.post(
        f"/api/v1/projects/{project['id']}/simulation-runs",
        json={"generation_run_id": run["id"], "input_schedule": {"1": {"SIG_TRAY_PRESENT": True}}, "max_cycles": 10, "expected_generation_revision": run["revision"]},
    )
    assert simulation.status_code == 201, simulation.text
    assert simulation.json()["verification_level"] == "automatic_reference"

    wrong_adapter = client.post(
        f"/api/v1/projects/{project['id']}/compile-runs",
        json={"generation_run_id": run["id"], "adapter_id": "gxworks3", "expected_generation_revision": run["revision"]},
    )
    assert wrong_adapter.status_code == 422
    assert wrong_adapter.json()["code"] == "COMPILE_ADAPTER_TARGET_MISMATCH"
    compile_run = client.post(
        f"/api/v1/projects/{project['id']}/compile-runs",
        json={"generation_run_id": run["id"], "adapter_id": "autoshop", "expected_generation_revision": run["revision"]},
    )
    assert compile_run.status_code == 201, compile_run.text
    assert compile_run.json()["status"] == "manual_required"
    assert compile_run.json()["verification_level"] == "unverified"

    matrix = client.get("/api/v1/compatibility-matrix")
    assert matrix.status_code == 200
    h5u = next(item for item in matrix.json()["entries"] if item["target"]["model"] == "H5U-1614MTD-A8")
    assert h5u["adapter_id"] == "autoshop"
    assert h5u["vendor_compile"] == "unverified"


def test_generation_requires_locked_spec(client: TestClient, project: dict) -> None:
    response = client.post(f"/api/v1/projects/{project['id']}/generation-runs", json={})
    assert response.status_code == 409
    assert response.json()["code"] == "LOCKED_SPEC_REQUIRED"


def test_deterministic_generator(locked_example: dict) -> None:
    spec = locked_example["revision"]["data"]
    first = generate_bundle(spec)
    second = generate_bundle(spec)
    assert first.files == second.files
    assert first.control_ir == second.control_ir
    assert first.test_spec == second.test_spec
    traced = {(item["entity_type"], item["entity_id"]) for item in first.trace_links}
    spec = locked_example["revision"]["data"]
    assert {("component", item["component_id"]) for item in spec["components"]} <= traced
    assert {("interlock", item["interlock_id"]) for item in spec["interlocks"]} <= traced
    assert {("exception", item["exception_id"]) for item in spec["exceptions"]} <= traced
    assert {("test_case", f"TEST_{item['step_id']}") for item in spec["sequence"]} <= traced


def test_generator_resolves_case_insensitive_signal_and_step_references(locked_example: dict) -> None:
    spec = locked_example["revision"]["data"]
    first_signal = spec["signals"][0]["signal_id"]
    next_step = spec["sequence"][1]["step_id"]
    spec["sequence"][0]["completion_condition"] = first_signal.lower()
    spec["sequence"][0]["next_step_id"] = next_step.lower()
    bundle = generate_bundle(spec)
    assert bundle.control_ir["steps"][0]["completion_condition"] == first_signal
    assert bundle.control_ir["steps"][0]["next_step_number"] == bundle.control_ir["steps"][1]["number"]


def test_generation_edit_commit_and_diff(
    client: TestClient,
    project: dict,
    locked_example: dict,
) -> None:
    generated = client.post(
        f"/api/v1/projects/{project['id']}/generation-runs",
        json={
            "spec_revision_id": locked_example["revision"]["id"],
            "branch_name": "generated/m1-acceptance",
        },
    )
    assert generated.status_code == 201, generated.text
    run = generated.json()
    assert run["status"] == "review_ready"
    assert {item["path"] for item in run["artifacts"]} >= {
        "src/GVL_IO.st",
        "src/PRG_AutoCycle.st",
        "generated/ControlIR.json",
        "tests/TestSpec.json",
    }
    assert run["trace_links"]

    branch_id = run["branch_id"]
    listing = client.get(f"/api/v1/branches/{branch_id}/files")
    assert listing.status_code == 200, listing.text
    branch = listing.json()["branch"]
    source = client.get(f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st")
    assert source.status_code == 200
    changed_content = source.json()["content"] + chr(10) + "// Reviewed locally." + chr(10)
    changed = client.patch(
        f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st",
        json={
            "content": changed_content,
            "reason": "工程师审阅标记",
            "expected_revision": branch["revision"],
        },
    )
    assert changed.status_code == 200, changed.text
    changed_branch = changed.json()["branch"]

    conflict = client.patch(
        f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st",
        json={
            "content": changed_content,
            "reason": "过期写入",
            "expected_revision": branch["revision"],
        },
    )
    assert conflict.status_code == 409

    committed = client.post(
        f"/api/v1/branches/{branch_id}/commits",
        json={
            "message": "Review generated auto cycle",
            "author": "测试工程师",
            "expected_revision": changed_branch["revision"],
        },
    )
    assert committed.status_code == 201, committed.text
    commit = committed.json()
    diff = client.get(f"/api/v1/commits/{commit['id']}/diff")
    assert diff.status_code == 200, diff.text
    assert "Reviewed locally" in diff.json()["diff"]
    commits = client.get(f"/api/v1/projects/{project['id']}/commits")
    assert len(commits.json()) == 2

    traversal = client.get(f"/api/v1/branches/{branch_id}/files/../secret.txt")
    assert traversal.status_code in {404, 422}


def test_concurrent_program_edits_recheck_revision_inside_repository_lock(
    client: TestClient,
    project: dict,
    locked_example: dict,
    monkeypatch,
) -> None:
    generated = client.post(
        f"/api/v1/projects/{project['id']}/generation-runs",
        json={
            "spec_revision_id": locked_example["revision"]["id"],
            "branch_name": "generated/concurrent-revision",
        },
    )
    assert generated.status_code == 201, generated.text
    branch_id = generated.json()["branch_id"]
    branch = client.get(f"/api/v1/branches/{branch_id}/files").json()["branch"]

    entered = threading.Event()
    release = threading.Event()
    original_write_files = api_main.write_files

    def blocking_write_files(repo, files):
        if not entered.is_set():
            entered.set()
            assert release.wait(timeout=5)
        return original_write_files(repo, files)

    monkeypatch.setattr(api_main, "write_files", blocking_write_files)
    payload = {
        "content": "// concurrent edit\n",
        "reason": "验证并发 revision 门禁",
        "expected_revision": branch["revision"],
    }
    endpoint = f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st"
    with ThreadPoolExecutor(max_workers=2) as executor:
        first = executor.submit(client.patch, endpoint, json=payload)
        assert entered.wait(timeout=5)
        second = executor.submit(client.patch, endpoint, json=payload)
        time.sleep(0.05)
        release.set()
        responses = [first.result(timeout=10), second.result(timeout=10)]

    assert sorted(response.status_code for response in responses) == [200, 409]
    conflict = next(response for response in responses if response.status_code == 409)
    assert conflict.json()["code"] == "REVISION_CONFLICT"


def test_each_program_commit_gets_independent_review_and_current_binding(
    client: TestClient,
    project: dict,
    locked_example: dict,
) -> None:
    generated = client.post(
        f"/api/v1/projects/{project['id']}/generation-runs",
        json={
            "spec_revision_id": locked_example["revision"]["id"],
            "branch_name": "generated/commit-review-binding",
        },
    )
    assert generated.status_code == 201, generated.text
    run = generated.json()
    branch_id = run["branch_id"]
    initial_reviews = client.get(
        f"/api/v1/projects/{project['id']}/automated-reviews"
    ).json()
    assert len(initial_reviews) == 1
    initial_commit_id = initial_reviews[0]["program_commit_id"]

    immutable = client.patch(
        f"/api/v1/branches/{branch_id}/files/generated/ControlIR.json",
        json={
            "content": "{}",
            "reason": "不应允许修改",
            "expected_revision": client.get(
                f"/api/v1/branches/{branch_id}/files"
            ).json()["branch"]["revision"],
        },
    )
    assert immutable.status_code == 409
    assert immutable.json()["code"] == "IMMUTABLE_GENERATION_BASELINE"

    source = client.get(
        f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st"
    ).json()["content"]
    branch = client.get(f"/api/v1/branches/{branch_id}/files").json()["branch"]
    edited = client.patch(
        f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st",
        json={
            "content": source + "\n// Deterministic local review note.\n",
            "reason": "增加不改变逻辑的审阅注释",
            "expected_revision": branch["revision"],
        },
    )
    assert edited.status_code == 200, edited.text
    legal_commit = client.post(
        f"/api/v1/branches/{branch_id}/commits",
        json={
            "message": "Add reviewed source comment",
            "author": "自动测试",
            "expected_revision": edited.json()["branch"]["revision"],
        },
    )
    assert legal_commit.status_code == 201, legal_commit.text
    legal_commit_id = legal_commit.json()["id"]
    assert legal_commit_id != initial_commit_id

    reviews = client.get(
        f"/api/v1/projects/{project['id']}/automated-reviews"
    ).json()
    legal_review = next(
        item for item in reviews if item["program_commit_id"] == legal_commit_id
    )
    assert legal_review["status"] == "passed"
    assert len(reviews) == 2

    current_run = client.get(f"/api/v1/generation-runs/{run['id']}").json()
    compile_run = client.post(
        f"/api/v1/projects/{project['id']}/compile-runs",
        json={
            "generation_run_id": run["id"],
            "adapter_id": "gxworks3",
            "expected_generation_revision": current_run["revision"],
        },
    )
    assert compile_run.status_code == 201, compile_run.text
    assert compile_run.json()["program_commit_id"] == legal_commit_id
    simulation = client.post(
        f"/api/v1/projects/{project['id']}/simulation-runs",
        json={
            "generation_run_id": run["id"],
            "expected_generation_revision": current_run["revision"],
        },
    )
    assert simulation.status_code == 201, simulation.text
    assert simulation.json()["program_commit_id"] == legal_commit_id

    branch = client.get(f"/api/v1/branches/{branch_id}/files").json()["branch"]
    dangerous = client.patch(
        f"/api/v1/branches/{branch_id}/files/src/PRG_AutoCycle.st",
        json={
            "content": source + "\nDOWNLOAD();\nKONGPU_UNDEFINED := TRUE;\n",
            "reason": "植入应被自动审核拦截的危险操作",
            "expected_revision": branch["revision"],
        },
    )
    assert dangerous.status_code == 200, dangerous.text
    blocked_commit = client.post(
        f"/api/v1/branches/{branch_id}/commits",
        json={
            "message": "Inject automatic-review regression defects",
            "author": "自动测试",
            "expected_revision": dangerous.json()["branch"]["revision"],
        },
    )
    assert blocked_commit.status_code == 201, blocked_commit.text
    blocked_commit_id = blocked_commit.json()["id"]
    blocked_review = next(
        item
        for item in client.get(
            f"/api/v1/projects/{project['id']}/automated-reviews"
        ).json()
        if item["program_commit_id"] == blocked_commit_id
    )
    assert blocked_review["status"] == "blocked"
    static_check = next(
        item
        for item in blocked_review["checks"]
        if item["id"] == "generation_static_audit"
    )
    assert static_check["status"] == "failed"

    current_run = client.get(f"/api/v1/generation-runs/{run['id']}").json()
    for endpoint, payload in (
        (
            f"/api/v1/projects/{project['id']}/compile-runs",
            {
                "generation_run_id": run["id"],
                "adapter_id": "gxworks3",
                "expected_generation_revision": current_run["revision"],
            },
        ),
        (
            f"/api/v1/projects/{project['id']}/simulation-runs",
            {
                "generation_run_id": run["id"],
                "expected_generation_revision": current_run["revision"],
            },
        ),
    ):
        rejected = client.post(endpoint, json=payload)
        assert rejected.status_code == 409
        assert rejected.json()["code"] == "AUTOMATED_REVIEW_BLOCKED"


def test_repository_path_guard(tmp_path) -> None:
    repo = tmp_path / "repo"
    repo.mkdir()
    try:
        safe_file(repo, "../escape.txt")
    except RepositoryError:
        pass
    else:
        raise AssertionError("path traversal should be rejected")


def test_identical_git_sha_is_scoped_to_program_branch(
    client: TestClient, project: dict
) -> None:
    second = client.post(
        "/api/v1/projects",
        json={"name": "第二个独立项目", "customer_code": "CUST-002"},
    )
    assert second.status_code == 201, second.text

    with client.app.state.database.session_factory() as session:
        first_workspace = ProgramWorkspace(
            project_id=project["id"], repository_path=project["id"]
        )
        second_workspace = ProgramWorkspace(
            project_id=second.json()["id"], repository_path=second.json()["id"]
        )
        session.add_all([first_workspace, second_workspace])
        session.flush()
        first_branch = ProgramBranch(
            workspace_id=first_workspace.id,
            name="generated/same-tree",
            git_ref="refs/heads/generated/same-tree",
        )
        second_branch = ProgramBranch(
            workspace_id=second_workspace.id,
            name="generated/same-tree",
            git_ref="refs/heads/generated/same-tree",
        )
        session.add_all([first_branch, second_branch])
        session.flush()
        shared_sha = "a" * 40
        session.add_all(
            [
                ProgramCommit(
                    branch_id=first_branch.id,
                    git_sha=shared_sha,
                    message="Same deterministic tree",
                ),
                ProgramCommit(
                    branch_id=second_branch.id,
                    git_sha=shared_sha,
                    message="Same deterministic tree",
                ),
            ]
        )
        session.commit()
        commits = session.scalars(
            select(ProgramCommit).where(ProgramCommit.git_sha == shared_sha)
        ).all()
        assert {item.branch_id for item in commits} == {
            first_branch.id,
            second_branch.id,
        }
