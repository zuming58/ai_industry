from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import PurePosixPath, PureWindowsPath
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, ConfigDict, Field

from .config import Settings
from .plc_profiles import TargetProfileError, profile_for_target


TEMPLATE_VERSION = "1.0"
SCHEMA_VERSION = "1.0"
REQUIRED_SHEETS = ["Instructions", "Project", "Components", "Signals", "Sequence"]
OPTIONAL_SHEETS = ["Interlocks", "Exceptions"]
ALL_SHEETS = REQUIRED_SHEETS + OPTIONAL_SHEETS

SHEET_COLUMNS: dict[str, list[str]] = {
    "Project": [
        "project_id", "project_name", "customer_code", "plc_brand", "plc_series",
        "plc_model", "software_version", "program_language", "cycle_target",
        "cycle_unit", "notes",
    ],
    "Components": [
        "component_id", "display_name", "parent_id", "component_type",
        "control_template", "parameter_value", "parameter_unit", "notes",
    ],
    "Signals": [
        "signal_id", "display_name", "direction", "address", "data_type",
        "unit", "component_id", "normal_state", "description",
    ],
    "Sequence": [
        "step_id", "display_name", "entry_condition", "actions",
        "completion_condition", "next_step_id", "expected_duration",
        "duration_unit", "parallel_group", "restart_policy",
    ],
    "Interlocks": [
        "interlock_id", "action_id", "allow_condition", "inhibit_condition",
        "external_safety_constraint", "description",
    ],
    "Exceptions": [
        "exception_id", "condition", "timeout_value", "timeout_unit",
        "response", "reset_condition", "operator_message",
    ],
}

ID_PATTERN = re.compile(r"^[A-Za-z][A-Za-z0-9_\-]{1,63}$")
SIGNAL_DIRECTIONS = {"DI", "DO", "AI", "AO", "INTERNAL", "COMM"}
DATA_TYPES = {"BOOL", "INT", "DINT", "REAL", "WORD", "DWORD", "STRING"}
BOOL_DIRECTIONS = {"DI", "DO"}
NUMERIC_DIRECTIONS = {"AI", "AO"}
WRITABLE_DIRECTIONS = {"DO", "AO", "INTERNAL", "COMM"}


def _fold_identifier(value: Any) -> str:
    """Return the IEC identifier comparison form without changing display text."""
    return str(value or "").strip().casefold()


class SourceRef(BaseModel):
    model_config = ConfigDict(extra="forbid")
    sheet: str
    row: int


class PlcTarget(BaseModel):
    brand: str
    series: str
    model: str
    software_version: str | None = None
    program_language: str = "ST"


class ProjectSpec(BaseModel):
    project_id: str
    project_name: str
    customer_code: str | None = None
    cycle_target: float | None = None
    cycle_unit: str | None = None
    notes: str | None = None
    source: SourceRef


class ComponentSpec(BaseModel):
    component_id: str
    display_name: str
    parent_id: str | None = None
    component_type: str
    control_template: str | None = None
    parameter_value: float | None = None
    parameter_unit: str | None = None
    notes: str | None = None
    source: SourceRef


class SignalSpec(BaseModel):
    signal_id: str
    display_name: str
    direction: str
    address: str | None = None
    data_type: str
    unit: str | None = None
    component_id: str | None = None
    normal_state: str | None = None
    description: str | None = None
    source: SourceRef


class SequenceStep(BaseModel):
    step_id: str
    display_name: str
    entry_condition: str | None = None
    actions: str
    completion_condition: str
    next_step_id: str | None = None
    expected_duration: float | None = None
    duration_unit: str | None = None
    parallel_group: str | None = None
    restart_policy: str | None = None
    source: SourceRef


class InterlockSpec(BaseModel):
    interlock_id: str
    action_id: str
    allow_condition: str
    inhibit_condition: str | None = None
    external_safety_constraint: str | None = None
    description: str | None = None
    source: SourceRef


class ExceptionSpec(BaseModel):
    exception_id: str
    condition: str
    timeout_value: float | None = None
    timeout_unit: str | None = None
    response: str
    reset_condition: str | None = None
    operator_message: str | None = None
    source: SourceRef


class MachineSpec(BaseModel):
    schema_version: str = SCHEMA_VERSION
    template_version: str = TEMPLATE_VERSION
    generated_at: str
    project: ProjectSpec
    plc_target: PlcTarget
    components: list[ComponentSpec] = Field(default_factory=list)
    signals: list[SignalSpec] = Field(default_factory=list)
    sequence: list[SequenceStep] = Field(default_factory=list)
    interlocks: list[InterlockSpec] = Field(default_factory=list)
    exceptions: list[ExceptionSpec] = Field(default_factory=list)


Severity = Literal["blocker", "warning", "suggestion", "info"]


@dataclass(frozen=True)
class IssueData:
    code: str
    severity: Severity
    title: str
    detail: str
    sheet: str | None = None
    row_number: int | None = None
    column_name: str | None = None
    entity_id: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class WorkbookInputError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def stable_json(data: dict[str, Any]) -> bytes:
    return json.dumps(data, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def spec_hash(data: dict[str, Any]) -> str:
    return hashlib.sha256(stable_json(data)).hexdigest()


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="0B55BB")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, start=1):
        max_length = max((len(str(cell.value or "")) for cell in column), default=10)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 3, 14), 34)


def _example_rows(project: dict[str, str]) -> dict[str, list[list[Any]]]:
    return {
        "Project": [[project["code"], project["name"], project.get("customer_code"), project["plc_brand"], project["plc_series"], project["plc_model"], "未检测", "ST", 18, "s", "脱敏演示项目"]],
        "Components": [
            ["STATION_01", "托盘举升检测站", None, "station", None, None, None, None],
            ["CYL_LIFT", "举升气缸", "STATION_01", "cylinder", "double_solenoid", 300, "mm", None],
            ["AXIS_TRANSFER", "移载轴", "STATION_01", "servo_axis", "axis_handshake", 425, "mm", None],
        ],
        "Signals": [
            ["SIG_TRAY_PRESENT", "托盘到位", "DI", "X010", "BOOL", None, "STATION_01", "FALSE", None],
            ["SIG_LIFT_EXTEND", "举升伸出", "DO", "Y020", "BOOL", None, "CYL_LIFT", "FALSE", None],
            ["SIG_LIFT_EXTENDED", "举升到位", "DI", "X011", "BOOL", None, "CYL_LIFT", "FALSE", None],
            ["SIG_AXIS_POSITION", "移载轴位置", "COMM", None, "REAL", "mm", "AXIS_TRANSFER", None, None],
        ],
        "Sequence": [
            ["S10", "等待托盘", "TRUE", "等待托盘", "SIG_TRAY_PRESENT", "S20", 1, "s", None, "resume"],
            ["S20", "举升定位", "SIG_TRAY_PRESENT", "SIG_LIFT_EXTEND := TRUE", "SIG_LIFT_EXTENDED", "S30", 3, "s", None, "restart_step"],
            ["S30", "移载取件", "SIG_LIFT_EXTENDED", "执行移载轴握手", "SIG_AXIS_POSITION >= 425", "END", 5, "s", None, "restart_step"],
        ],
        "Interlocks": [["ILK_LIFT", "SIG_LIFT_EXTEND", "SIG_TRAY_PRESENT", "AxisMoving", "安全回路由外部安全系统负责", None]],
        "Exceptions": [["EXC_LIFT_TIMEOUT", "NOT SIG_LIFT_EXTENDED", 3, "s", "停止当前工步并提示", "ResetCommand", "举升未到位"]],
    }


def generate_workbook(project: dict[str, str], kind: str = "blank") -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    instructions = wb.create_sheet("Instructions")
    instructions.append(["控谱 MachineSpec Excel Template", f"Template v{TEMPLATE_VERSION}"])
    instructions.append(["填写顺序", "Project → Components → Signals → Sequence → 选填项"])
    instructions.append(["规则", "不得修改工作表和列名；稳定 ID 与显示名称分离；未知值留空并说明。"])
    instructions.append(["安全边界", "安全 PLC、急停、门锁等安全功能不由本模板自动生成。"])
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 90

    examples = _example_rows(project) if kind == "example" else {}
    for sheet_name in SHEET_COLUMNS:
        ws = wb.create_sheet(sheet_name)
        ws.append(SHEET_COLUMNS[sheet_name])
        for row in examples.get(sheet_name, []):
            ws.append(row)
        _style_header(ws)

    direction_validation = DataValidation(type="list", formula1='"DI,DO,AI,AO,INTERNAL,COMM"')
    datatype_validation = DataValidation(type="list", formula1='"BOOL,INT,DINT,REAL,WORD,DWORD,STRING"')
    signals = wb["Signals"]
    signals.add_data_validation(direction_validation)
    signals.add_data_validation(datatype_validation)
    direction_validation.add("C2:C2000")
    datatype_validation.add("E2:E2000")

    meta = wb.create_sheet("_meta")
    meta.sheet_state = "hidden"
    for key, value in [
        ("template_version", TEMPLATE_VERSION),
        ("schema_version", SCHEMA_VERSION),
        ("project_id", project["id"]),
        ("project_code", project["code"]),
        ("plc_brand", project["plc_brand"]),
        ("plc_series", project["plc_series"]),
        ("plc_model", project["plc_model"]),
        ("generated_at", datetime.now(timezone.utc).isoformat()),
        ("kind", kind),
    ]:
        meta.append([key, value])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def inspect_xlsx_archive(content: bytes, settings: Settings) -> None:
    if len(content) > settings.max_upload_bytes:
        raise WorkbookInputError("FILE_TOO_LARGE", "文件超过 20 MB 上限")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > settings.max_xlsx_entries:
                raise WorkbookInputError("XLSX_TOO_MANY_ENTRIES", "工作簿内部文件数量异常")
            total_size = sum(entry.file_size for entry in entries)
            if total_size > settings.max_xlsx_uncompressed_bytes:
                raise WorkbookInputError("XLSX_EXPANSION_LIMIT", "工作簿解压后体积异常")
            normalized_names: list[str] = []
            for entry in entries:
                normalized = entry.filename.replace("\\", "/")
                posix_path = PurePosixPath(normalized)
                windows_path = PureWindowsPath(normalized)
                if (
                    not normalized
                    or normalized.startswith("/")
                    or posix_path.is_absolute()
                    or windows_path.is_absolute()
                    or bool(windows_path.drive)
                    or any(part in {"", ".", ".."} for part in normalized.split("/"))
                ):
                    raise WorkbookInputError("XLSX_PATH_TRAVERSAL", "工作簿内部路径不安全")
                if entry.flag_bits & 0x1:
                    raise WorkbookInputError("XLSX_ENCRYPTED", "不接受加密的 XLSX 工作簿")
                lowered = normalized.lower()
                if (
                    lowered.endswith("vbaproject.bin")
                    or lowered.startswith("xl/activex/")
                    or lowered.startswith("xl/ctrlprops/")
                ):
                    raise WorkbookInputError("XLSX_ACTIVE_CONTENT", "不接受宏或 ActiveX 内容")
                normalized_names.append(normalized)
            if len(normalized_names) != len(set(normalized_names)):
                raise WorkbookInputError("XLSX_DUPLICATE_ENTRY", "工作簿包含重复的内部路径")
            if "[Content_Types].xml" not in set(normalized_names):
                raise WorkbookInputError("INVALID_XLSX", "文件不是有效的 XLSX 工作簿")
    except zipfile.BadZipFile as exc:
        raise WorkbookInputError("INVALID_XLSX", "文件损坏或不是有效的 XLSX 工作簿") from exc


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _rows(ws, expected_columns: list[str]) -> tuple[list[dict[str, Any]], list[IssueData]]:
    headers = [_clean(cell.value) for cell in ws[1]]
    issues: list[IssueData] = []
    missing = [column for column in expected_columns if column not in headers]
    for column in missing:
        issues.append(IssueData("MISSING_COLUMN", "blocker", "缺少必填列", f"{ws.title} 缺少列 {column}", ws.title, 1, column))
    index = {name: position for position, name in enumerate(headers) if name}
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {column: _clean(values[index[column]]) if column in index and index[column] < len(values) else None for column in expected_columns}
        if not any(value is not None for value in data.values()):
            continue
        data["source"] = {"sheet": ws.title, "row": row_number}
        rows.append(data)
    return rows, issues


def parse_workbook(content: bytes, settings: Settings) -> tuple[dict[str, Any], list[IssueData]]:
    inspect_xlsx_archive(content, settings)
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False, keep_vba=False)
    except Exception as exc:
        raise WorkbookInputError("WORKBOOK_OPEN_FAILED", "工作簿无法读取，可能已加密或损坏") from exc

    issues: list[IssueData] = []
    for sheet in REQUIRED_SHEETS:
        if sheet not in workbook.sheetnames:
            issues.append(IssueData("MISSING_SHEET", "blocker", "缺少必填工作表", f"缺少工作表 {sheet}", sheet))
    meta_values: dict[str, Any] = {}
    if "_meta" in workbook.sheetnames:
        meta_values = {str(row[0]): row[1] for row in workbook["_meta"].iter_rows(values_only=True) if row and row[0]}
    else:
        issues.append(IssueData("MISSING_META", "blocker", "缺少模板元数据", "工作簿不是由当前项目模板生成", "_meta"))

    parsed: dict[str, list[dict[str, Any]]] = {}
    for sheet_name, columns in SHEET_COLUMNS.items():
        if sheet_name in workbook.sheetnames:
            parsed[sheet_name], row_issues = _rows(workbook[sheet_name], columns)
            issues.extend(row_issues)
        else:
            parsed[sheet_name] = []

    project_rows = parsed["Project"]
    if not project_rows:
        issues.append(IssueData("EMPTY_PROJECT", "blocker", "项目表为空", "Project 必须包含一行项目资料", "Project", 2))
        project_row = {key: None for key in SHEET_COLUMNS["Project"]}
        project_row["source"] = {"sheet": "Project", "row": 2}
    else:
        project_row = project_rows[0]
        if len(project_rows) > 1:
            issues.append(IssueData("MULTIPLE_PROJECT_ROWS", "warning", "Project 包含多行", "仅第一行作为项目元数据", "Project", 3))

    spec = {
        "schema_version": str(meta_values.get("schema_version") or SCHEMA_VERSION),
        "template_version": str(meta_values.get("template_version") or TEMPLATE_VERSION),
        "_meta": {
            "schema_version": str(meta_values.get("schema_version") or ""),
            "template_version": str(meta_values.get("template_version") or ""),
            "project_id": str(meta_values.get("project_id") or ""),
        },
        "generated_at": str(meta_values.get("generated_at") or datetime.now(timezone.utc).isoformat()),
        "project": {
            "project_id": str(project_row.get("project_id") or ""),
            "project_name": str(project_row.get("project_name") or ""),
            "customer_code": project_row.get("customer_code"),
            "cycle_target": project_row.get("cycle_target"),
            "cycle_unit": project_row.get("cycle_unit"),
            "notes": project_row.get("notes"),
            "source": project_row["source"],
        },
        "plc_target": {
            "brand": str(project_row.get("plc_brand") or ""),
            "series": str(project_row.get("plc_series") or ""),
            "model": str(project_row.get("plc_model") or ""),
            "software_version": project_row.get("software_version"),
            "program_language": str(project_row.get("program_language") or "ST"),
        },
        "components": parsed["Components"],
        "signals": parsed["Signals"],
        "sequence": parsed["Sequence"],
        "interlocks": parsed["Interlocks"],
        "exceptions": parsed["Exceptions"],
    }
    return spec, issues


def _identifier_issues(items: list[dict[str, Any]], key: str, sheet: str) -> list[IssueData]:
    issues: list[IssueData] = []
    seen: set[str] = set()
    for item in items:
        value = str(item.get(key) or "")
        source = item["source"]
        if not value:
            issues.append(IssueData("MISSING_ID", "blocker", "稳定 ID 为空", f"{key} 为必填项", sheet, source["row"], key))
        elif not ID_PATTERN.match(value):
            issues.append(IssueData("INVALID_ID", "blocker", "稳定 ID 格式无效", f"{value} 必须以字母开头且只能包含字母、数字、下划线或短横线", sheet, source["row"], key, value))
        elif _fold_identifier(value) in seen:
            issues.append(IssueData("DUPLICATE_ID", "blocker", "稳定 ID 重复", f"{value} 在 {sheet} 中重复", sheet, source["row"], key, value))
        seen.add(_fold_identifier(value))
    return issues


def _references(expression: str | None, candidates: set[str]) -> set[str]:
    if not expression:
        return set()
    tokens = set(re.findall(r"[A-Za-z][A-Za-z0-9_\-]*", expression))
    keywords = {"TRUE", "FALSE", "AND", "OR", "NOT", "END"}
    folded_candidates = {_fold_identifier(item) for item in candidates}
    return {token for token in tokens if _fold_identifier(token) not in folded_candidates and token.upper() not in keywords and not token.isdigit()}


def _restricted_expression(expression: Any) -> bool:
    return expression is None or bool(
        re.fullmatch(r"[A-Za-z0-9_\- \t.()+*/<>=]+", str(expression))
        and not any(token in str(expression) for token in ("//", "/*", "*/", "(*", "*)"))
    )


def validate_spec(spec: dict[str, Any], expected_project: dict[str, str] | None = None) -> list[IssueData]:
    issues: list[IssueData] = []
    project = spec["project"]
    plc = spec["plc_target"]
    metadata = spec.get("_meta") or {}
    if spec.get("schema_version") != SCHEMA_VERSION:
        issues.append(IssueData("SCHEMA_VERSION_UNSUPPORTED", "blocker", "Schema 版本不受支持", f"收到 {spec.get('schema_version') or '空值'}，当前支持 {SCHEMA_VERSION}", "_meta", 1, "schema_version"))
    if spec.get("template_version") != TEMPLATE_VERSION:
        issues.append(IssueData("TEMPLATE_VERSION_UNSUPPORTED", "blocker", "模板版本不受支持", f"收到 {spec.get('template_version') or '空值'}，当前支持 {TEMPLATE_VERSION}", "_meta", 1, "template_version"))
    if metadata.get("schema_version") and metadata.get("schema_version") != spec.get("schema_version"):
        issues.append(IssueData("META_SCHEMA_MISMATCH", "blocker", "模板元数据版本不一致", "_meta 与解析出的 Schema 版本不一致", "_meta", 1, "schema_version"))
    if metadata.get("template_version") and metadata.get("template_version") != spec.get("template_version"):
        issues.append(IssueData("META_TEMPLATE_MISMATCH", "blocker", "模板元数据版本不一致", "_meta 与解析出的模板版本不一致", "_meta", 1, "template_version"))
    target_profile = None
    try:
        target_profile = profile_for_target(plc)
    except TargetProfileError as exc:
        issues.append(IssueData("PLC_TARGET_UNSUPPORTED", "blocker", "PLC 目标不受支持", str(exc), "Project", project["source"]["row"], "plc_model"))
    if not project.get("project_name"):
        issues.append(IssueData("PROJECT_NAME_REQUIRED", "blocker", "项目名称为空", "Project.project_name 为必填项", "Project", project["source"]["row"], "project_name"))
    if not project.get("project_id"):
        issues.append(IssueData("PROJECT_ID_REQUIRED", "blocker", "项目 ID 为空", "Project.project_id 为必填项", "Project", project["source"]["row"], "project_id"))
    if project.get("cycle_target") is not None and not project.get("cycle_unit"):
        issues.append(IssueData("UNIT_REQUIRED", "blocker", "节拍缺少单位", "cycle_target 有值时必须填写 cycle_unit", "Project", project["source"]["row"], "cycle_unit"))
    if expected_project:
        expected = (expected_project["plc_brand"], expected_project["plc_series"], expected_project["plc_model"])
        actual = (plc.get("brand"), plc.get("series"), plc.get("model"))
        if expected != actual:
            issues.append(IssueData("PLC_TARGET_MISMATCH", "blocker", "PLC 目标与项目不一致", f"模板目标为 {actual}，项目目标为 {expected}", "Project", project["source"]["row"], "plc_model"))
        if project.get("project_id") not in {expected_project["id"], expected_project["code"]}:
            issues.append(IssueData("PROJECT_ID_MISMATCH", "blocker", "模板项目与当前项目不一致", "请从当前项目重新下载模板", "Project", project["source"]["row"], "project_id"))
        meta_project_id = metadata.get("project_id")
        if meta_project_id and meta_project_id not in {expected_project["id"], expected_project["code"]}:
            issues.append(IssueData("META_PROJECT_ID_MISMATCH", "blocker", "模板元数据与当前项目不一致", "请从当前项目重新下载模板", "_meta", 1, "project_id"))

    issues.extend(_identifier_issues(spec["components"], "component_id", "Components"))
    issues.extend(_identifier_issues(spec["signals"], "signal_id", "Signals"))
    issues.extend(_identifier_issues(spec["sequence"], "step_id", "Sequence"))
    issues.extend(_identifier_issues(spec["interlocks"], "interlock_id", "Interlocks"))
    issues.extend(_identifier_issues(spec["exceptions"], "exception_id", "Exceptions"))

    for collection, sheet, message in (
        (spec["components"], "Components", "至少定义一个设备或元件"),
        (spec["signals"], "Signals", "至少定义一个信号"),
        (spec["sequence"], "Sequence", "至少定义一个流程步骤"),
    ):
        if not collection:
            issues.append(IssueData("REQUIRED_DATA_EMPTY", "blocker", f"{sheet} 没有数据", message, sheet, 2))

    component_ids = {_fold_identifier(item["component_id"]) for item in spec["components"] if item.get("component_id")}
    signal_ids = {_fold_identifier(item["signal_id"]) for item in spec["signals"] if item.get("signal_id")}
    step_ids = {_fold_identifier(item["step_id"]) for item in spec["sequence"] if item.get("step_id")}

    for component in spec["components"]:
        source = component["source"]
        for column in ("display_name", "component_type"):
            if not component.get(column):
                issues.append(IssueData("REQUIRED_FIELD_MISSING", "blocker", "元件必填字段为空", f"Components.{column} 为必填项", "Components", source["row"], column, component.get("component_id")))
        parent = component.get("parent_id")
        if parent and _fold_identifier(parent) not in component_ids:
            source = component["source"]
            issues.append(IssueData("COMPONENT_PARENT_MISSING", "blocker", "父级元件不存在", f"{parent} 未在 Components 中定义", "Components", source["row"], "parent_id", component.get("component_id")))
        if component.get("parameter_value") is not None and not component.get("parameter_unit"):
            source = component["source"]
            issues.append(IssueData("UNIT_REQUIRED", "blocker", "元件参数缺少单位", "parameter_value 有值时必须填写 parameter_unit", "Components", source["row"], "parameter_unit", component.get("component_id")))

    address_owner: dict[str, str] = {}
    for signal in spec["signals"]:
        source = signal["source"]
        direction = str(signal.get("direction") or "").upper()
        data_type = str(signal.get("data_type") or "").upper()
        signal_id = signal.get("signal_id")
        for column in ("display_name", "direction", "data_type"):
            if not signal.get(column):
                issues.append(IssueData("REQUIRED_FIELD_MISSING", "blocker", "信号必填字段为空", f"Signals.{column} 为必填项", "Signals", source["row"], column, signal_id))
        if direction not in SIGNAL_DIRECTIONS:
            issues.append(IssueData("INVALID_SIGNAL_DIRECTION", "blocker", "信号方向无效", f"{direction} 不在允许枚举中", "Signals", source["row"], "direction", signal_id))
        if data_type not in DATA_TYPES:
            issues.append(IssueData("INVALID_DATA_TYPE", "blocker", "数据类型无效", f"{data_type} 不在允许枚举中", "Signals", source["row"], "data_type", signal_id))
        if direction in BOOL_DIRECTIONS and data_type and data_type != "BOOL":
            issues.append(IssueData("SIGNAL_TYPE_MISMATCH", "blocker", "信号方向与数据类型不匹配", f"{direction} 首版必须使用 BOOL", "Signals", source["row"], "data_type", signal_id))
        if direction in NUMERIC_DIRECTIONS and data_type == "BOOL":
            issues.append(IssueData("SIGNAL_TYPE_MISMATCH", "blocker", "模拟量不能使用 BOOL", f"{direction} 应使用数值类型", "Signals", source["row"], "data_type", signal_id))
        if data_type in {"INT", "DINT", "REAL"} and not signal.get("unit"):
            issues.append(IssueData("UNIT_REQUIRED", "warning", "数值信号缺少单位", "数值信号应明确工程单位", "Signals", source["row"], "unit", signal_id))
        component_id = signal.get("component_id")
        if component_id and _fold_identifier(component_id) not in component_ids:
            issues.append(IssueData("SIGNAL_COMPONENT_MISSING", "blocker", "信号引用的元件不存在", f"{component_id} 未在 Components 中定义", "Signals", source["row"], "component_id", signal_id))
        address = str(signal.get("address") or "").upper()
        if direction in {"DI", "DO"} and not address:
            issues.append(IssueData("IO_ADDRESS_REQUIRED", "blocker", "数字量信号地址为空", f"{direction} 信号必须填写 I/O 地址", "Signals", source["row"], "address", signal_id))
        if address:
            if target_profile and not target_profile.address_pattern.fullmatch(address):
                issues.append(IssueData("INVALID_IO_ADDRESS", "blocker", "I/O 地址格式不受目标 Profile 支持", f"{address} 不符合 {target_profile.profile_id} 的首批 X/Y/M 逻辑地址子集", "Signals", source["row"], "address", signal_id))
            expected_prefix = {"DI": "X", "DO": "Y", "INTERNAL": "M"}.get(direction)
            if expected_prefix and not address.startswith(expected_prefix):
                issues.append(IssueData("IO_DIRECTION_ADDRESS_MISMATCH", "blocker", "信号方向与地址类型不匹配", f"{direction} 信号应使用 {expected_prefix} 地址，收到 {address}", "Signals", source["row"], "address", signal_id))
            if address in address_owner:
                issues.append(IssueData("DUPLICATE_IO_ADDRESS", "blocker", "I/O 地址冲突", f"{address} 已被 {address_owner[address]} 使用", "Signals", source["row"], "address", signal_id))
            address_owner[address] = str(signal_id)

    terminal_steps = {None, "", "END"}
    edges: dict[str, str | None] = {}
    for step in spec["sequence"]:
        source = step["source"]
        step_id = step.get("step_id")
        for column in ("display_name", "actions", "completion_condition"):
            if not step.get(column):
                issues.append(IssueData("REQUIRED_FIELD_MISSING", "blocker", "流程必填字段为空", f"Sequence.{column} 为必填项", "Sequence", source["row"], column, step_id))
        next_step = step.get("next_step_id")
        folded_step_id = _fold_identifier(step_id)
        folded_next_step = _fold_identifier(next_step)
        edges[folded_step_id] = folded_next_step if next_step else None
        if next_step not in terminal_steps and folded_next_step != "end" and folded_next_step not in step_ids:
            issues.append(IssueData("NEXT_STEP_MISSING", "blocker", "下一步骤不存在", f"{next_step} 未在 Sequence 中定义", "Sequence", source["row"], "next_step_id", step_id))
        if step.get("expected_duration") is not None and not step.get("duration_unit"):
            issues.append(IssueData("UNIT_REQUIRED", "blocker", "步骤时长缺少单位", "expected_duration 有值时必须填写 duration_unit", "Sequence", source["row"], "duration_unit", step_id))
        for column in ("entry_condition", "actions", "completion_condition"):
            for unknown in _references(step.get(column), signal_ids):
                issues.append(IssueData("SIGNAL_REFERENCE_MISSING", "blocker", "步骤引用的信号不存在", f"{unknown} 未在 Signals 中定义", "Sequence", source["row"], column, step_id))
        for column in ("entry_condition", "completion_condition"):
            if not _restricted_expression(step.get(column)):
                issues.append(IssueData("EXPRESSION_SYNTAX_UNSUPPORTED", "blocker", "表达式超出受限子集", f"Sequence.{column} 包含不允许的字符或语法", "Sequence", source["row"], column, step_id))
        for statement in (value.strip() for value in re.split(r"[;\r\n]+", str(step.get("actions") or "")) if value.strip()):
            assignment = re.fullmatch(r"([A-Za-z][A-Za-z0-9_-]*)\s*:=\s*(.+)", statement)
            if assignment:
                target_signal = next(
                    (item for item in spec["signals"] if _fold_identifier(item.get("signal_id")) == _fold_identifier(assignment.group(1))),
                    None,
                )
                if target_signal and str(target_signal.get("direction") or "").upper() not in WRITABLE_DIRECTIONS:
                    issues.append(IssueData("ACTION_TARGET_DIRECTION_INVALID", "blocker", "动作目标不是可写信号", f"{assignment.group(1)} 不能由 Sequence 动作写入", "Sequence", source["row"], "actions", step_id))
                if not _restricted_expression(assignment.group(2)):
                    issues.append(IssueData("EXPRESSION_SYNTAX_UNSUPPORTED", "blocker", "动作表达式超出受限子集", "赋值右侧包含不允许的字符或语法", "Sequence", source["row"], "actions", step_id))
            else:
                issues.append(IssueData("ACTION_NOT_DETERMINISTIC", "warning", "动作不能确定性生成", f"“{statement}”不是首版 信号 := 表达式 子集，将只保留 TODO", "Sequence", source["row"], "actions", step_id))

    signal_by_id = {
        _fold_identifier(item.get("signal_id")): item for item in spec["signals"]
        if item.get("signal_id")
    }
    for interlock in spec["interlocks"]:
        source = interlock["source"]
        action_id = interlock.get("action_id")
        if not action_id:
            issues.append(IssueData("INTERLOCK_ACTION_REQUIRED", "blocker", "互锁动作为空", "Interlocks.action_id 为必填项", "Interlocks", source["row"], "action_id", interlock.get("interlock_id")))
        elif _fold_identifier(action_id) not in signal_by_id:
            issues.append(IssueData("INTERLOCK_ACTION_MISSING", "blocker", "互锁动作信号不存在", f"{action_id} 未在 Signals 中定义", "Interlocks", source["row"], "action_id", interlock.get("interlock_id")))
        elif str(signal_by_id[_fold_identifier(action_id)].get("direction") or "").upper() not in WRITABLE_DIRECTIONS:
            issues.append(IssueData("INTERLOCK_ACTION_NOT_WRITABLE", "blocker", "互锁动作不是可写信号", f"{action_id} 不能作为互锁控制动作", "Interlocks", source["row"], "action_id", interlock.get("interlock_id")))
        if not interlock.get("allow_condition"):
            issues.append(IssueData("INTERLOCK_ALLOW_REQUIRED", "blocker", "互锁允许条件为空", "Interlocks.allow_condition 为必填项", "Interlocks", source["row"], "allow_condition", interlock.get("interlock_id")))
        for column in ("allow_condition", "inhibit_condition"):
            if not _restricted_expression(interlock.get(column)):
                issues.append(IssueData("EXPRESSION_SYNTAX_UNSUPPORTED", "blocker", "表达式超出受限子集", f"Interlocks.{column} 包含不允许的字符或语法", "Interlocks", source["row"], column, interlock.get("interlock_id")))
            for unknown in _references(interlock.get(column), signal_ids):
                issues.append(IssueData("INTERLOCK_EXTERNAL_STATE_UNVERIFIED", "warning", "互锁引用外部状态", f"{unknown} 未在 Signals 中定义，自动验证只能按外部只读状态处理", "Interlocks", source["row"], column, interlock.get("interlock_id")))

    for exception in spec["exceptions"]:
        source = exception["source"]
        for column in ("condition", "response"):
            if not exception.get(column):
                issues.append(IssueData("REQUIRED_FIELD_MISSING", "blocker", "异常必填字段为空", f"Exceptions.{column} 为必填项", "Exceptions", source["row"], column, exception.get("exception_id")))
        if exception.get("timeout_value") is not None and not exception.get("timeout_unit"):
            issues.append(IssueData("UNIT_REQUIRED", "blocker", "异常超时缺少单位", "timeout_value 有值时必须填写 timeout_unit", "Exceptions", source["row"], "timeout_unit", exception.get("exception_id")))
        for column in ("condition", "reset_condition"):
            if not _restricted_expression(exception.get(column)):
                issues.append(IssueData("EXPRESSION_SYNTAX_UNSUPPORTED", "blocker", "表达式超出受限子集", f"Exceptions.{column} 包含不允许的字符或语法", "Exceptions", source["row"], column, exception.get("exception_id")))
            for unknown in _references(exception.get(column), signal_ids):
                issues.append(IssueData("EXCEPTION_EXTERNAL_STATE_UNVERIFIED", "warning", "异常引用外部状态", f"{unknown} 未在 Signals 中定义，需在厂商工程中核对来源", "Exceptions", source["row"], column, exception.get("exception_id")))
        issues.append(IssueData("EXCEPTION_VENDOR_LOGIC_REQUIRED", "warning", "异常逻辑尚未生成到 ST", "当前确定性生成器只将异常写入 Control IR 与 TestSpec；计时器、报警输出和复位逻辑需在厂商工程集中验证", "Exceptions", source["row"], "response", exception.get("exception_id")))

    if spec["sequence"]:
        start = _fold_identifier(spec["sequence"][0].get("step_id"))
        reachable: set[str] = set()
        current: str | None = start
        while current and current not in reachable and current != "end":
            reachable.add(current)
            current = edges.get(current)
        for step in spec["sequence"]:
            step_id = _fold_identifier(step.get("step_id"))
            if step_id not in reachable:
                source = step["source"]
                issues.append(IssueData("UNREACHABLE_STEP", "blocker", "流程步骤不可达", f"{step_id} 无法从首步骤到达", "Sequence", source["row"], "step_id", step_id))
        if current in reachable:
            cycle_nodes: set[str] = set()
            node = current
            while node and node not in cycle_nodes:
                cycle_nodes.add(node)
                node = edges.get(node)
            exits = [edges.get(node) for node in cycle_nodes if edges.get(node) not in cycle_nodes]
            if not exits:
                issues.append(IssueData("FLOW_CYCLE_WITHOUT_EXIT", "blocker", "流程循环没有退出路径", f"循环包含：{', '.join(sorted(cycle_nodes))}", "Sequence"))

    if not spec["exceptions"]:
        issues.append(IssueData("EXCEPTIONS_EMPTY", "suggestion", "异常表为空", "允许继续，但建议补充已知超时、异常和复位策略", "Exceptions"))
    return issues


SHEET_ENTITY_MAP = {
    "Project": ("project", None),
    "Components": ("components", "component_id"),
    "Signals": ("signals", "signal_id"),
    "Sequence": ("sequence", "step_id"),
    "Interlocks": ("interlocks", "interlock_id"),
    "Exceptions": ("exceptions", "exception_id"),
}


def patch_cells(spec: dict[str, Any], edits: list[dict[str, Any]]) -> dict[str, Any]:
    updated = json.loads(json.dumps(spec, ensure_ascii=False))
    for edit in edits:
        sheet = edit["sheet"]
        row = int(edit["row"])
        column = edit["column"]
        if sheet not in SHEET_ENTITY_MAP or column not in SHEET_COLUMNS[sheet]:
            raise WorkbookInputError("INVALID_CELL_TARGET", f"不能编辑 {sheet}.{column}")
        collection_name, _id_key = SHEET_ENTITY_MAP[sheet]
        collection = updated[collection_name]
        targets = [collection] if sheet == "Project" else collection
        target = next((item for item in targets if item.get("source", {}).get("row") == row), None)
        if target is None:
            raise WorkbookInputError("CELL_NOT_FOUND", f"{sheet} 第 {row} 行不存在")
        target[column] = _clean(edit.get("value"))
    return updated


def required_review_views(spec: dict[str, Any]) -> list[str]:
    views = ["device_relationship", "process_flow", "cycle_analysis", "signal_timing", "io_mapping", "raw_tables"]
    if spec.get("interlocks"):
        views.append("interlock_matrix")
    if spec.get("exceptions"):
        views.append("exceptions")
    return views


def sheet_payload(spec: dict[str, Any], sheet: str) -> dict[str, Any]:
    if sheet not in SHEET_ENTITY_MAP:
        raise WorkbookInputError("UNKNOWN_SHEET", f"未知工作表 {sheet}")
    collection_name, _id_key = SHEET_ENTITY_MAP[sheet]
    data = spec[collection_name]
    rows = [data] if sheet == "Project" else data
    return {"sheet": sheet, "columns": SHEET_COLUMNS[sheet], "rows": rows}
